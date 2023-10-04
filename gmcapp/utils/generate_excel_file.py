from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse, JsonResponse

def generate_excel_file(table_id, selected_records, filename):
    # Create an HttpResponse containing the Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

    # Create a new Excel workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Define column headers based on the table_id
    if table_id == 'web-data-table':
        headers = ['SNO.', 'Title', 'Website Link']
    elif table_id == 'facebook-data-table':
        headers = ['SNO.', 'Title', 'Page Link', 'Source']
    else:
        return JsonResponse({'success': False, 'error_message': 'Invalid table_id'})

    # Write headers to the first row
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet[f'{col_letter}1']
        cell.value = header

    # Write data rows
    for row_num, record in enumerate(selected_records, 2):
        worksheet[f'A{row_num}'] = row_num - 1  # SNO.
        if table_id == 'web-data-table':
            worksheet[f'B{row_num}'] = record.title
            worksheet[f'C{row_num}'] = record.website_link
        elif table_id == 'facebook-data-table':
            worksheet[f'B{row_num}'] = record.title
            worksheet[f'C{row_num}'] = record.page_link
            worksheet[f'D{row_num}'] = record.source

    # Save the workbook to the response
    workbook.save(response)
    return response

# Example usage:
# For web-data-table
# response = generate_excel_file('web-data-table', selected_web_records, 'selected_web_records')

# For facebook-data-table
# response = generate_excel_file('facebook-data-table', selected_facebook_records, 'selected_facebook_records')
